import openpyxl

""""Worbook class to handle worksheet use in the bot"""
class Workbook:
    
    """"Load or create the workbook and spreadsheet"""
    def __init__(self):
        self.number_of_columns = 4
        self.dest_filename = "Jobs.xlsx"
        
        try:
            self.wb = openpyxl.load_workbook("Jobs.xlsx")
            self.ws = self.wb["sheet1"]
            
            print("File succesfully loaded")
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "sheet1"
            
            index = 0
            self.header = ["Title: ", "Company: ", "Location", "Link: "]
            
            for column in range (1, self.number_of_columns+1):
                self.ws.cell(1, column, self.header[index])
                index += 1
                
            empty_row = ["", "", "", ""]
            self.write_row(empty_row)
            
            print("File succesfully created")

    """Handle row insertion in the workbook"""
    def write_row(self, info):
        inputrow = len(self.ws["A"]) + 1
        index = 0
        
        for column in range (1, self.number_of_columns+1):
            self.ws.cell(inputrow, column, info[index])
            index += 1

    """Save the workbook with the new data"""
    def save(self):
        self.wb.save(self.dest_filename)
